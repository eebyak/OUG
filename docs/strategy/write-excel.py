"""
Open University of Germany Financial Model Generator
Version 1.1
Generates professional Excel file from JSON specification
"""
import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime
#cell.number_format = '#,##0.00'


    
def load_json_config(filename="data.json"):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(base_dir, filename)

    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

def create_inputs_sheet(wb, config):
    """Create the Inputs sheet with all assumptions"""
    ws = wb.create_sheet("Inputs", 0)
    
    # Title
    ws['A1'] = "OUG FINANCIAL MODEL - INPUT PARAMETERS v1.1"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Formatting styles
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=11, color="1F4E79")
    number_format = '#,##0'
    currency_format = '#,##0 €'
    
    row = 3
    
    # A. GROWTH & SCALE
    ws[f'A{row}'] = "A. GROWTH & SCALE ASSUMPTIONS"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    growth = config['assumptions']['growth']
    growth_items = [
        ("Start students Year 1", growth['start_students_year_1'], "Realistic first cohort"),
        ("Annual growth rate", growth['annual_growth_rate'], "Conservative"),
        ("Second program launch year", growth['second_program_launch_year'], "Adds 25% boost"),
        ("Second program boost", growth['second_program_growth_boost'], growth['second_program_boost_description']),
        ("ECTS per degree student/year", growth['ects_per_degree_student'], "Part-time average"),
        ("Microcredential learners Year 1", growth['microcredential_start_year_1'], "Early adopters"),
        ("Microcredential growth rate", growth['microcredential_growth_rate'], "Faster initially"),
        ("ECTS per microcredential learner", growth['ects_per_microcredential'], "Single module"),
        ("Dropout rate", growth['dropout_rate'], "Applied to partner payouts")
    ]
    
    for label, value, note in growth_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        if isinstance(value, (int, float)):
            ws[f'B{row}'].number_format = '0.00' if isinstance(value, float) and value < 1 else '#,##0'
        row += 1
    
    row += 1
    
    # B. PRICING & REVENUE
    ws[f'A{row}'] = "B. PRICING & REVENUE"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    pricing = config['assumptions']['pricing']
    pricing_items = [
        ("Price per ECTS (EUR)", pricing['price_per_ects'], "Fixed pricing"),
        ("State funding starts (Year)", pricing['state_funding_start_year'], "After public status"),
        ("State funding per student (EUR)", pricing['state_funding_per_student'], "Conservative estimate"),
        ("Drittmittel annual (EUR)", pricing['drittmittel_per_year'], "EU/DAAD/BMBF"),
        ("Industry sponsorship annual (EUR)", pricing['sponsorship_per_year'], "Board-level support")
    ]
    
    for label, value, note in pricing_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        if isinstance(value, (int, float)):
            ws[f'B{row}'].number_format = currency_format if value > 100 else '#,##0'
        row += 1
    
    row += 1
    
    # C. PARTNER PAYOUTS
    ws[f'A{row}'] = "C. PARTNER PAYOUTS"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    partner = config['assumptions']['partner_payouts']
    partner_items = [
        ("Partner share of tuition", partner['partner_share'], "Teaching compensation"),
        ("OUG overhead retained", partner['oug_overhead'], "Central operations")
    ]
    
    for label, value, note in partner_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        ws[f'B{row}'].number_format = '0.00'
        row += 1
    
    row += 1
    
    # D. PERSONNEL COSTS
    ws[f'A{row}'] = "D. PERSONNEL COSTS (CORE TEAM)"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Personnel table headers
    headers = ["Position", "Tarif", "VZÄ", "Annual Cost (EUR)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="E6F0FA")
    row += 1
    
    # Personnel data
    for person in config['assumptions']['personnel']['base_team']:
        ws[f'A{row}'] = person['position']
        ws[f'B{row}'] = person['tarif']
        ws[f'C{row}'] = person['vza']
        ws[f'D{row}'] = person['annual_cost']
        ws[f'D{row}'].number_format = currency_format
        row += 1
    
    # Total
    ws[f'A{row}'] = "TOTAL PERSONNEL YEAR 1"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = config['assumptions']['personnel']['base_team_total']
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = currency_format
    row += 2
    
    # E. OPERATING EXPENSES
    ws[f'A{row}'] = "E. OPERATING EXPENSES (OPEX)"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # OPEX table
    ws[f'A{row}'] = "Category/Sub-item"
    ws[f'B{row}'] = "Annual Cost (EUR)"
    ws[f'C{row}'] = "Scaling"
    ws[f'D{row}'] = "Notes"
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="E6F0FA")
    row += 1
    
    # OPEX categories and items
    for category in config['assumptions']['opex']['categories']:
        # Category header
        ws[f'A{row}'] = category['name']
        ws[f'A{row}'].font = Font(bold=True, italic=True)
        ws[f'B{row}'] = category['base_cost']
        ws[f'B{row}'].number_format = currency_format
        row += 1
        
        # Items
        for item in category['items']:
            ws[f'A{row}'] = f"  {item['name']}"
            ws[f'B{row}'] = item['cost']
            ws[f'C{row}'] = item['scaling']
            ws[f'D{row}'] = item.get('description', '')
            ws[f'B{row}'].number_format = currency_format
            row += 1
    
    # Total OPEX
    ws[f'A{row}'] = "TOTAL BASE OPEX YEAR 1"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = config['assumptions']['opex']['base_year_1']
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "Year 0 pre-launch: " + str(config['assumptions']['opex']['year_0_prelaunch'])
    row += 2
    
    # F. MARKETING BUDGET
    ws[f'A{row}'] = "F. MARKETING BUDGET BY YEAR"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for item in config['assumptions']['marketing']['budget_by_year']:
        ws[f'A{row}'] = f"Marketing Year {item['year']}"
        ws[f'B{row}'] = item['budget']
        ws[f'C{row}'] = item['phase']
        ws[f'B{row}'].number_format = currency_format
        row += 1
    
    row += 1
    
    # G. CAPEX
    ws[f'A{row}'] = "G. CAPEX (ONE-TIME INVESTMENTS)"
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    for block in config['assumptions']['capex']['blocks']:
        ws[f'A{row}'] = block['name']
        ws[f'B{row}'] = block['total']
        ws[f'B{row}'].number_format = currency_format
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for spend in block['spend_by_year']:
            if spend['year'] <= 1:  # Only show Years 0-1
                ws[f'A{row}'] = f"  {block['name']} - Year {spend['year']}"
                ws[f'B{row}'] = spend['amount']
                ws[f'C{row}'] = ', '.join(spend.get('items', [])[:2]) + "..."
                ws[f'B{row}'].number_format = currency_format
                row += 1
    
    # Totals
    ws[f'A{row}'] = "TOTAL CAPEX"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = config['assumptions']['capex']['total']
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    row += 1
    
    ws[f'A{row}'] = "Amortization period (years)"
    ws[f'B{row}'] = config['assumptions']['capex']['amortization_years']
    row += 1
    
    ws[f'A{row}'] = "Annual Amortization"
    ws[f'B{row}'] = config['assumptions']['capex']['annual_amortization']
    ws[f'B{row}'].number_format = currency_format
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30 if col == 'A' else 15

def create_calculations_sheet(wb, config):
    """Create the Calculations sheet with 10-year projection"""
    ws = wb.create_sheet("Calculations", 1)
    
    # Title
    ws['A1'] = "OUG FINANCIAL MODEL - 10-YEAR PROJECTION"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:K1')
    
    # Get yearly data from outputs
    yearly_data = config['outputs']['yearly_table']
    
    # Headers
    headers = ["Description"] + [f"Year {d['year']}" for d in yearly_data]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6F0FA")
        cell.alignment = Alignment(horizontal="center")
    
    # Define rows to display (ALWAYS: (desc, key, calc_func))
    rows_to_show = [
        ("Degree Students", "degree_students", None),
        ("Microcredential Learners", "microcredential_learners", None),

        ("ECTS from Degree", None, lambda d: d["degree_students"] * config["assumptions"]["growth"]["ects_per_degree_student"]),
        ("ECTS from Micro", None, lambda d: d["microcredential_learners"] * config["assumptions"]["growth"]["ects_per_microcredential"]),
        ("Total ECTS Sold", "total_ects", None),
        ("ECTS Completed", None, lambda d: d["total_ects"] * (1 - config["assumptions"]["growth"]["dropout_rate"])),

        ("", None, None),
        ("REVENUE", None, None),

        ("Tuition Revenue", None, lambda d: d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"]),
        #("State Funding", None, lambda d: d.get("total_revenue_state", 0)),
        #("State Funding", None, lambda d:
        #    d["total_revenue"]
        #    - (d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"])
        #    - (config["assumptions"]["pricing"]["drittmittel_per_year"]
        #    if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0)
        #    - (config["assumptions"]["pricing"]["sponsorship_per_year"]
        #    if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0)
        #),
        ("State Funding", None, lambda d: max(0,
            d["total_revenue"]
            - (d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"])
            - (config["assumptions"]["pricing"]["drittmittel_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0)
            - (config["assumptions"]["pricing"]["sponsorship_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0)
        )),



        ("Drittmittel", None, lambda d: config["assumptions"]["pricing"]["drittmittel_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0),
        ("Industry Sponsorship", None, lambda d: config["assumptions"]["pricing"]["sponsorship_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0),


        ("Other / rounding", None, lambda d:
            d["total_revenue"]
            - max(0,
                d["total_revenue"]
                - (d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"])
                - (config["assumptions"]["pricing"]["drittmittel_per_year"]
                if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0)
                - (config["assumptions"]["pricing"]["sponsorship_per_year"]
                if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0)
            )
            - (d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"])
            - (config["assumptions"]["pricing"]["drittmittel_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0)
            - (config["assumptions"]["pricing"]["sponsorship_per_year"]
            if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0)
        ),



        ("Total Revenue", "total_revenue", None),


        #("Tuition Revenue", "total_revenue", lambda d: d["total_ects"] * config["assumptions"]["pricing"]["price_per_ects"]),
        #("State Funding", None, lambda d: d["total_revenue_state"] if "total_revenue_state" in d else 0),
        #("Drittmittel", None, lambda d: config["assumptions"]["pricing"]["drittmittel_per_year"]
        #    if d["year"] >= config["assumptions"]["pricing"]["drittmittel_start_year"] else 0),
        #("Industry Sponsorship", None, lambda d: config["assumptions"]["pricing"]["sponsorship_per_year"]
        #    if d["year"] >= config["assumptions"]["pricing"]["sponsorship_start_year"] else 0),
        #("Total Revenue", "total_revenue", None),

        ("", None, None),
        ("EXPENSES", None, None),

        ("Partner Payouts", "partner_payouts", None),
        ("Personnel Costs", "personnel_costs", None),
        ("OPEX", "opex_costs", None),
        ("Marketing", "marketing_costs", None),
        ("Total Operating Expenses", "total_expenses", None),
        ("Operating Result", "operating_result", None),

        ("", None, None),

        ("CAPEX Spend", None, lambda d:
            config["assumptions"]["capex"]["blocks"][0]["spend_by_year"][0]["amount"] if d["year"] == 0 else
            (config["assumptions"]["capex"]["blocks"][0]["spend_by_year"][1]["amount"] if d["year"] == 1 else 0)
        ),
        ("Amortization", None, lambda d: config["assumptions"]["capex"]["annual_amortization"] if d["year"] >= 1 else 0),

        ("", None, None),

        ("Cashflow", "cashflow", None),
        ("Cumulative Cashflow", "cumulative_cashflow", None),
    ]

    # Fill data
    for row_idx, (desc, key, calc_func) in enumerate(rows_to_show, start=4):
        ws.cell(row=row_idx, column=1).value = desc
        if desc:
            ws.cell(row=row_idx, column=1).font = Font(bold=True) if desc in ["REVENUE", "EXPENSES"] else Font()
        
        for col_idx, year_data in enumerate(yearly_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if calc_func is not None:
                cell.value = calc_func(year_data)
            elif key and key in year_data:
                cell.value = year_data[key]

            #if key and key in year_data:
            #    cell.value = year_data[key]
            #elif calc_func:
            #    cell.value = calc_func(year_data)
            
            # Formatting
            if isinstance(cell.value, (int, float)):
                if desc in ["Operating Result", "Cashflow", "Cumulative Cashflow"]:
                    cell.number_format = '#,##0 €'
                    # Color negative numbers red
                    if cell.value < 0:
                        cell.font = Font(color="FF0000")
                else:
                    cell.number_format = '#,##0'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(4, 4 + len(rows_to_show)):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 12):
        ws.column_dimensions[chr(64 + col)].width = 15

def create_summary_sheet(wb, config):
    """Create the Summary sheet with key metrics and insights"""
    ws = wb.create_sheet("Summary", 2)
    
    # Title
    ws['A1'] = "OPEN UNIVERSITY OF GERMANY - FINANCIAL SUMMARY v1.1"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws.merge_cells('A1:F1')
    ws['A2'] = f"Generated: {datetime.now().strftime('%d.%m.%Y')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:F2')
    
    # Key Metrics Box
    ws['A4'] = "KEY METRICS"
    ws['A4'].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells('A4:F4')
    
    metrics = [
        ("Total Start-up Capital Required", "EUR 7,100,000", "CAPEX €4.1M + Operating support €3.0M"),
        ("Peak Funding Need", "EUR 11,585,797", "Year 7 (maximum cumulative deficit)"),
        ("Operational Break-even", "Year 8", "Operating result turns positive"),
        ("Students at Break-even", "974", "Degree students"),
        ("Students at Year 5", "732", "Degree students + 350 micro = 1,082 total"),
        ("Students at Year 10", "1,179", "Degree students + 704 micro = 1,883 total"),
        ("Total ECTS Delivered Year 10", "38,887", "Equivalent to 648 full-time student years"),
        ("Cumulative Partner Payouts by Year 5", "EUR 883,000", "Real money to teaching universities"),
        ("Cumulative Partner Payouts by Year 10", "EUR 2.4M", "Growing returns to partners"),
        ("Cost per Student at Scale", "EUR 2,250", "vs. €8-10k at traditional universities")
    ]
    
    row = 5
    for metric, value, note in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, color="1F4E79")
        row += 1
    
    # Year-by-Year Results Table
    row += 2
    ws[f'A{row}'] = "YEAR-BY-YEAR RESULTS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Table headers
    headers = ["Year", "Students", "Revenue (EUR)", "Expenses (EUR)", "Result (EUR)", "Cumulative (EUR)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6F0FA")
        cell.alignment = Alignment(horizontal="center")
    row += 1
    
    # Table data
    for year_data in config['outputs']['yearly_table']:
        ws[f'A{row}'] = f"Year {year_data['year']}"
        ws[f'B{row}'] = year_data['degree_students']
        ws[f'C{row}'] = year_data['total_revenue']
        ws[f'D{row}'] = year_data['total_expenses']
        ws[f'E{row}'] = year_data['operating_result']
        ws[f'F{row}'] = year_data['cumulative_cashflow']
        
        # Format numbers
        for col in ['C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].number_format = '#,##0 €'
        
        # Color negative results red
        if year_data['operating_result'] < 0:
            ws[f'E{row}'].font = Font(color="FF0000")
        if year_data['cumulative_cashflow'] < 0:
            ws[f'F{row}'].font = Font(color="FF0000")
        
        row += 1
    
    # Key Insights
    row += 2
    ws[f'A{row}'] = "KEY INSIGHTS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    insights = [
        "1. Total funding need: EUR 7.1M (CAPEX EUR 4.1M + Operating support EUR 3.0M over first 4 years)",
        "2. Break-even (operational): Year 8 when revenue covers all operating expenses",
        "3. Cost efficiency: EUR 2,250 per student at scale vs EUR 8-10k at traditional universities",
        "4. Partner payouts return 60% of tuition to teaching universities",
        "5. State funding (Grundmittel) begins Year 5 after public university status achieved",
        "6. Model is resilient: even at 8% growth, break-ever occurs by Year 9"
    ]
    
    for insight in insights:
        ws[f'A{row}'] = insight
        row += 1
    
    # Strategic Case
    row += 2
    ws[f'A{row}'] = "STRATEGIC CASE"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4E79")
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    strategic_text = [
        "Private Fernhochschulen already enroll hundreds of thousands of students who cannot access public universities.",
        "These students pay €2,000-4,000 per year for education that should be a public good.",
        "",
        "With €7.1M public investment, the OUG offers quality-assured, stackable credentials at €600/semester—",
        "1/3 to 1/6 the cost of private alternatives—while returning 60% of tuition to public universities."
    ]
    
    for line in strategic_text:
        ws[f'A{row}'] = line
        if line:
            ws[f'A{row}'].font = Font(italic=True) if line.startswith("With") else Font()
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25

def create_scenarios_sheet(wb, config):
    """Create optional Scenarios sheet with sensitivity analysis"""
    ws = wb.create_sheet("Scenarios", 3)
    
    ws['A1'] = "SCENARIO ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:E1')
    
    row = 3
    
    # Scenario comparison table
    headers = ["Scenario", "Growth Rate", "Price/ECTS", "State Funding", "Break-even Year", "Peak Funding"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E6F0FA")
    row += 1
    
    # Data for each scenario
    scenarios = config['assumptions']['scenarios']
    
    # Conservative
    ws[f'A{row}'] = scenarios['conservative']['name']
    ws[f'B{row}'] = scenarios['conservative']['growth_rate']
    ws[f'C{row}'] = scenarios['conservative']['price_per_ects']
    ws[f'D{row}'] = scenarios['conservative']['state_funding_per_student']
    ws[f'E{row}'] = "Year 9-10"
    ws[f'F{row}'] = "EUR 13.2M"
    row += 1
    
    # Realistic
    ws[f'A{row}'] = scenarios['realistic']['name']
    ws[f'B{row}'] = scenarios['realistic']['growth_rate']
    ws[f'C{row}'] = scenarios['realistic']['price_per_ects']
    ws[f'D{row}'] = scenarios['realistic']['state_funding_per_student']
    ws[f'E{row}'] = "Year 8"
    ws[f'F{row}'] = "EUR 11.6M"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Optimistic
    ws[f'A{row}'] = scenarios['optimistic']['name']
    ws[f'B{row}'] = scenarios['optimistic']['growth_rate']
    ws[f'C{row}'] = scenarios['optimistic']['price_per_ects']
    ws[f'D{row}'] = scenarios['optimistic']['state_funding_per_student']
    ws[f'E{row}'] = "Year 6-7"
    ws[f'F{row}'] = "EUR 9.8M"
    row += 3
    
    # Format numbers
    for r in range(row-3, row):
        for col in ['B', 'C', 'D']:
            ws[f'{col}{r}'].number_format = '0.00' if col == 'B' else '#,##0'
    
    # Sensitivity analysis note
    ws[f'A{row}'] = "SENSITIVITY INSIGHTS:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    insights = [
        "• The model is most sensitive to student growth rate and state funding per student",
        "• A 2% lower growth rate delays break-even by approximately 2 years",
        "• Partner share of 60% creates strong incentives; reducing to 55% in optimistic scenario accelerates break-even",
        "• Marketing investment in Years 1-3 is critical to achieving growth assumptions"
    ]
    
    for insight in insights:
        ws[f'A{row}'] = insight
        row += 1
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

def generate_excel_file(config, filename="OUG_Financial_Model_v1.1.xlsx"):
    """Main function to generate the complete Excel file"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create all sheets
    create_inputs_sheet(wb, config)
    create_calculations_sheet(wb, config)
    create_summary_sheet(wb, config)
    create_scenarios_sheet(wb, config)
    
    # Save the file
    wb.save(filename)
    print(f"✅ Excel file generated successfully: {filename}")
    return filename

# Main execution
if __name__ == "__main__":
    # Load configuration
    # In practice, you would load from file
    # config = load_json_config()
    
    # For this example, we'll note that the JSON should be pasted
    print("To use this script:")
    print("1. Paste the complete JSON from Part 1 into the load_json_config function")
    print("2. Run: python oug_financial_model_generator.py")
    print("\nOr modify the script to load from an external JSON file:")
    print("   with open('oug_financial_model_v1.1.json', 'r') as f:")
    print("       config = json.load(f)")
    
    # Uncomment to actually generate:

    #print("Script location:", os.path.abspath(__file__))
    #print("Working directory:", os.getcwd())
    #print("Files in working directory:", os.listdir())

    config = load_json_config()
    print("Top-level keys:", config.keys())

    generate_excel_file(config)